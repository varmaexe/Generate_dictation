"""
generate_dictation.py
Generates an Excel file with N blocks of the Dictation structure.

Usage:
    python generate_dictation.py --blocks 4
    python generate_dictation.py --blocks 10 --output my_file.xlsx
    python generate_dictation.py --blocks 4 --min 1 --max 9
    python generate_dictation.py --blocks 4 --sheets 3       # spread across multiple sheets
    python generate_dictation.py --blocks 4 --double         # odd rows=single digit, even rows=double digit
    python generate_dictation.py --blocks 4 --double --min 2 --max 8  # custom single-digit range
    python generate_dictation.py --blocks 4 --all-double     # all rows=double digit (10–99)
    python generate_dictation.py --blocks 4 --double-triple  # odd rows=double digit, even rows=triple digit
"""

import os
import re
import random
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Config ──────────────────────────────────────────────────────────────────

DATA_ROWS    = 5          # rows per group (A, B, C)
NUM_GROUPS   = 3          # groups per block (A, B, C)
NUM_COLS     = 5          # data columns (1–5)
BLOCK_GAP    = 1          # empty columns between blocks
HEADER_ROWS  = 1          # header row at top of block

GROUP_LABELS = ["A", "B", "C "]

SUMMARY_DEFS = [
    ("A",    1,  5),
    ("B",    6,  10),
    ("C ",   11, 15),
    ("A-C",  1,  15),
    ("A-6",  1,  6),
    ("A-7",  1,  7),
    ("A-8",  1,  8),
    ("A-9",  1,  9),
    ("AB",   1,  10),
    ("BC",   6,  15),
    ("B-6",  6,  11),
    ("B-7",  6,  12),
    ("B-8",  6,  13),
    ("B-9",  6,  14),
]

# Excel colors
COLOR_HEADER_BG   = "4472C4"
COLOR_HEADER_FG   = "FFFFFF"
COLOR_GROUP_A     = "DDEEFF"
COLOR_GROUP_B     = "EEFFDD"
COLOR_GROUP_C     = "FFEECC"
COLOR_GROUP_A_ALT = "BBCCE8"
COLOR_GROUP_B_ALT = "CCE8BB"
COLOR_GROUP_C_ALT = "E8DDBB"
COLOR_SUMMARY_BG  = "F2F2F2"
COLOR_LABEL_BG    = "D9D9D9"
COLOR_GAP         = "EBEBEB"

# Terminal colors
TP  = "\033[35m"    # magenta/purple
TPL = "\033[95m"    # light magenta
TC  = "\033[36m"    # cyan
TY  = "\033[93m"    # yellow
TW  = "\033[97m"    # bright white
TG  = "\033[92m"    # green
TR  = "\033[91m"    # red
TRS = "\033[0m"     # reset
TD  = "\033[2m"     # dim


# ── Terminal helpers ─────────────────────────────────────────────────────────

def section(title: str):
    line = "─" * (44 - len(title) - 3)
    print(f"\n  {TD}── {title} {line}{TRS}")


def box_row(label: str, value, width: int = 24) -> str:
    value_str = str(value)
    pad = width - len(value_str)
    return f"  │  {TD}{label:<13}{TRS}: {TC}{value_str}{TRS}{' ' * pad}│"


# ── Interactive prompt helpers ───────────────────────────────────────────────

def ask_int(question: str, default: int, min_val: int = None, max_val: int = None) -> int:
    while True:
        raw = input(f"  {TW}{question}{TRS} {TD}(default: {TY}{default}{TD}){TRS}: ").strip()
        if raw == "":
            return default
        try:
            val = int(raw)
            if min_val is not None and val < min_val:
                print(f"    {TY}⚠{TRS}  Please enter a number >= {min_val}")
                continue
            if max_val is not None and val > max_val:
                print(f"    {TY}⚠{TRS}  Please enter a number <= {max_val}")
                continue
            return val
        except ValueError:
            print(f"    {TY}⚠{TRS}  Please enter a whole number")


def ask_yes_no(question: str, default: bool) -> bool:
    default_hint = "Y/n" if default else "y/N"
    while True:
        raw = input(f"  {TW}{question}{TRS} {TD}({default_hint}){TRS}: ").strip().lower()
        if raw == "":
            return default
        if raw in ("y", "yes"):
            return True
        if raw in ("n", "no"):
            return False
        print(f"    {TY}⚠{TRS}  Please enter y or n")


def ask_str(question: str, default: str) -> str:
    raw = input(f"  {TW}{question}{TRS} {TD}(default: {TY}{default}{TD}){TRS}: ").strip()
    return raw if raw else default


# ── Filename helper ──────────────────────────────────────────────────────────

def sanitize_filename(name: str) -> str:
    """Replace characters invalid in filenames across platforms."""
    # Strip extension first to sanitize the base name only
    base = name[:-5] if name.lower().endswith(".xlsx") else name
    base = re.sub(r'[\\/:*?"<>|\s]+', '_', base).strip('_')
    return (base or "dictation") + ".xlsx"


# ── Excel helpers ────────────────────────────────────────────────────────────

def col_letter(col: int) -> str:
    return get_column_letter(col)


def make_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)


def thin_border() -> Border:
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)


def write_block(ws, block_index: int, start_col: int, val_min: int, val_max: int, double_mode: bool = False, all_double: bool = False, all_triple: bool = False, double_triple: bool = False):
    """
    Write one full block at the given starting column.
    Block layout (rows):
      Row 1        : Header  ("Block N", 1, 2, 3, 4, 5)
      Rows 2–6     : Group A data
      Rows 7–11    : Group B data
      Rows 12–16   : Group C data
      Rows 17–30   : Summary formulas
    """

    DATA_START_ROW  = 2
    SUMMARY_START_R = DATA_START_ROW + NUM_GROUPS * DATA_ROWS
    TOTAL_ROWS      = 1 + NUM_GROUPS * DATA_ROWS + len(SUMMARY_DEFS)

    label_col      = start_col
    data_col_start = start_col + 1
    gap_col        = data_col_start + NUM_COLS

    # ── Header row ──────────────────────────────────────────────────────────
    hdr_fill  = make_fill(COLOR_HEADER_BG)
    hdr_font  = Font(bold=True, color=COLOR_HEADER_FG, name="Arial", size=10)
    hdr_align = Alignment(horizontal="center")

    ws.row_dimensions[1].height = 20

    cell = ws.cell(row=1, column=label_col, value=f"Block {block_index + 1}")
    cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = hdr_align; cell.border = thin_border()

    for i, col_num in enumerate(range(1, NUM_COLS + 1)):
        cell = ws.cell(row=1, column=data_col_start + i, value=col_num)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = hdr_align; cell.border = thin_border()

    # ── Data rows (Groups A, B, C) ───────────────────────────────────────────
    group_colors     = [COLOR_GROUP_A,     COLOR_GROUP_B,     COLOR_GROUP_C]
    alt_group_colors = [COLOR_GROUP_A_ALT, COLOR_GROUP_B_ALT, COLOR_GROUP_C_ALT]
    is_alternating   = double_mode or double_triple

    for g_idx, (g_label, g_color) in enumerate(zip(GROUP_LABELS, group_colors)):
        base_fill  = make_fill(g_color)
        alt_fill   = make_fill(alt_group_colors[g_idx])
        label_font = Font(bold=True, name="Arial", size=10)
        data_font  = Font(name="Arial", size=10)
        center     = Alignment(horizontal="center")

        for r in range(DATA_ROWS):
            excel_row   = DATA_START_ROW + g_idx * DATA_ROWS + r
            is_even_row = (r + 1) % 2 == 0
            row_fill    = alt_fill if (is_alternating and is_even_row) else base_fill

            ws.row_dimensions[excel_row].height = 18

            lbl_cell        = ws.cell(row=excel_row, column=label_col)
            lbl_cell.fill   = row_fill
            lbl_cell.border = thin_border()
            if r == 0:
                lbl_cell.value     = g_label
                lbl_cell.font      = label_font
                lbl_cell.alignment = Alignment(horizontal="left")

            for c in range(NUM_COLS):
                if all_triple:
                    val = random.randint(100, 999)
                elif all_double:
                    val = random.randint(10, 99)
                elif double_triple and is_even_row:
                    val = random.randint(100, 999)
                elif double_triple:
                    val = random.randint(10, 99)
                elif double_mode and is_even_row:
                    val = random.randint(10, 99)
                else:
                    val = random.randint(val_min, val_max)
                cell           = ws.cell(row=excel_row, column=data_col_start + c, value=val)
                cell.font      = data_font
                cell.fill      = row_fill
                cell.alignment = center
                cell.border    = thin_border()

    # ── Summary rows ─────────────────────────────────────────────────────────
    sum_label_font = Font(bold=True, name="Arial", size=10)
    sum_data_font  = Font(name="Arial", size=10)
    sum_fill       = make_fill(COLOR_SUMMARY_BG)
    lbl_fill       = make_fill(COLOR_LABEL_BG)
    center         = Alignment(horizontal="center")

    for s_idx, (s_label, rel_start, rel_end) in enumerate(SUMMARY_DEFS):
        excel_row = SUMMARY_START_R + s_idx
        abs_start = DATA_START_ROW + rel_start - 1
        abs_end   = DATA_START_ROW + rel_end   - 1

        ws.row_dimensions[excel_row].height = 16

        lbl           = ws.cell(row=excel_row, column=label_col, value=s_label)
        lbl.font      = sum_label_font
        lbl.fill      = lbl_fill
        lbl.alignment = Alignment(horizontal="left")
        lbl.border    = thin_border()

        for c in range(NUM_COLS):
            excel_col      = data_col_start + c
            col_ltr        = col_letter(excel_col)
            formula        = f"=SUM({col_ltr}{abs_start}:{col_ltr}{abs_end})"
            cell           = ws.cell(row=excel_row, column=excel_col, value=formula)
            cell.font      = sum_data_font
            cell.fill      = sum_fill
            cell.alignment = center
            cell.border    = thin_border()

    # ── Gap column ────────────────────────────────────────────────────────────
    gap_fill = make_fill(COLOR_GAP)
    for row in range(1, TOTAL_ROWS + 1):
        ws.cell(row=row, column=gap_col).fill = gap_fill

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions[col_letter(label_col)].width  = 9
    ws.column_dimensions[col_letter(gap_col)].width    = 2
    for c in range(NUM_COLS):
        ws.column_dimensions[col_letter(data_col_start + c)].width = 6


# ── Main ────────────────────────────────────────────────────────────────────

def run_once():
    """Collect settings, generate one file, return True to loop again."""

    # ── Setup ────────────────────────────────────────────────────────────────
    section("Setup")
    blocks = ask_int("How many blocks do you want?", default=4, min_val=1)
    sheets = ask_int("How many sheets to spread blocks across?", default=1, min_val=1, max_val=blocks)

    # ── Mode ─────────────────────────────────────────────────────────────────
    section("Mode")
    print(f"    {TY}1{TRS}  All single digit   {TD}(1–9){TRS}")
    print(f"    {TY}2{TRS}  All double digit   {TD}(10–99){TRS}")
    print(f"    {TY}3{TRS}  All triple digit   {TD}(100–999){TRS}")
    print(f"    {TY}4{TRS}  Alternating        {TD}(odd rows = single digit, even rows = double digit){TRS}")
    print(f"    {TY}5{TRS}  Alternating        {TD}(odd rows = double digit, even rows = triple digit){TRS}")
    print()
    mode_choice = ask_int("Select mode", default=1, min_val=1, max_val=5)

    double        = (mode_choice == 4)
    all_double    = (mode_choice == 2)
    all_triple    = (mode_choice == 3)
    double_triple = (mode_choice == 5)

    mode_labels = {
        1: f"All single digit  {TD}(1–9){TRS}",
        2: f"All double digit  {TD}(10–99){TRS}",
        3: f"All triple digit  {TD}(100–999){TRS}",
        4: f"Alternating       {TD}(single + double){TRS}",
        5: f"Alternating       {TD}(double + triple){TRS}",
    }
    print(f"  {TG}✓{TRS}  {mode_labels[mode_choice]}")

    # ── Range ────────────────────────────────────────────────────────────────
    if double:
        section("Range")
        print(f"  {TD}Single-digit range applies to odd rows only. Double-digit is always 10–99.{TRS}")
        val_min = ask_int("Min value for single-digit rows", default=1, min_val=1, max_val=9)
        val_max = ask_int("Max value for single-digit rows", default=9, min_val=val_min, max_val=9)
    elif all_double:
        val_min, val_max = 10, 99
    elif all_triple:
        val_min, val_max = 100, 999
    elif double_triple:
        val_min, val_max = 10, 999
    else:
        section("Range")
        val_min = ask_int("Min random value", default=1, min_val=1)
        val_max = ask_int("Max random value", default=9, min_val=val_min)

    # ── Output ───────────────────────────────────────────────────────────────
    section("Output")
    default_seed = random.randint(10000, 99999)
    seed = ask_int("Random seed (reuse to reproduce this exact file)", default=default_seed, min_val=0)

    now          = datetime.now().strftime("%Y-%m-%d_%H%M")
    raw_output   = ask_str("Output filename", default=f"dictation_{now}_seed{seed}.xlsx")
    output       = sanitize_filename(raw_output)
    if output != sanitize_filename(raw_output.strip()):
        pass  # sanitize_filename is deterministic; just use the result
    if output != raw_output and not raw_output.lower().endswith(".xlsx"):
        output = sanitize_filename(raw_output)
    # Show warning if the name was changed
    cleaned = sanitize_filename(raw_output)
    if cleaned != raw_output and cleaned != raw_output + ".xlsx":
        print(f"  {TY}⚠{TRS}  Filename sanitized to: {TC}{cleaned}{TRS}")
    output = cleaned

    if os.path.exists(output):
        print(f"  {TY}⚠{TRS}  '{output}' already exists.")
        if not ask_yes_no("Overwrite?", default=False):
            print(f"\n  {TD}Cancelled. No file was created.{TRS}\n")
            return

    # ── Summary ───────────────────────────────────────────────────────────────
    print()
    print(f"  ┌─────────────────────────────────────────┐")
    print(f"  │  {TW}Summary{TRS}                                │")
    print(f"  ├─────────────────────────────────────────┤")
    print(box_row("Blocks", blocks))
    print(box_row("Sheets", sheets))
    if double:
        print(box_row("Mode", "Alternating (single + double)"))
        print(box_row("Single range", f"{val_min}–{val_max}"))
        print(box_row("Double range", "10–99"))
    elif all_double:
        print(box_row("Mode", "All double digit"))
        print(box_row("Value range", "10–99"))
    elif all_triple:
        print(box_row("Mode", "All triple digit"))
        print(box_row("Value range", "100–999"))
    elif double_triple:
        print(box_row("Mode", "Alternating (double + triple)"))
        print(box_row("Double range", "10–99"))
        print(box_row("Triple range", "100–999"))
    else:
        print(box_row("Mode", "All single digit"))
        print(box_row("Value range", f"{val_min}–{val_max}"))
    print(box_row("Output file", output))
    print(box_row("Seed", seed))
    print(f"  └─────────────────────────────────────────┘")
    print()

    if not ask_yes_no("Generate file with these settings?", default=True):
        print(f"\n  {TD}Cancelled. No file was created.{TRS}\n")
        return

    # ── Generate ─────────────────────────────────────────────────────────────
    print()
    random.seed(seed)

    wb = Workbook()
    wb.remove(wb.active)

    blocks_per_sheet = (blocks + sheets - 1) // sheets
    block_width      = 1 + NUM_COLS + BLOCK_GAP
    last_row         = 1 + NUM_GROUPS * DATA_ROWS + len(SUMMARY_DEFS)

    block_index = 0
    for sheet_num in range(sheets):
        print(f"  {TD}→{TRS} Writing sheet {TW}{sheet_num + 1}{TRS} of {sheets} ...", end="", flush=True)
        ws = wb.create_sheet(title=f"Dictation {sheet_num + 1}")
        blocks_this_sheet = min(blocks_per_sheet, blocks - block_index)

        for b in range(blocks_this_sheet):
            start_col = 1 + b * block_width
            write_block(ws, block_index, start_col, val_min, val_max, double, all_double, all_triple, double_triple)
            block_index += 1

        last_col = 1 + (blocks_this_sheet - 1) * block_width + NUM_COLS
        ws.print_area = f"A1:{col_letter(last_col)}{last_row}"

        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToPage   = True
        ws.page_setup.fitToWidth  = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True

        ws.freeze_panes = "A2"

        # Seed metadata note (below all content, small grey text)
        meta_row  = last_row + 2
        meta_cell = ws.cell(row=meta_row, column=1,
                            value=f"Seed: {seed}  ·  Generated: {now}")
        meta_cell.font = Font(name="Arial", size=8, color="AAAAAA", italic=True)

        print(f"  {TG}✓{TRS}")

    wb.save(output)

    # ── Done ─────────────────────────────────────────────────────────────────
    full_path  = os.path.abspath(output)
    info_line  = f"{blocks} block(s) · {sheets} sheet(s) · seed {seed}"
    w1         = max(len(output), len(info_line), len(full_path)) + 2
    bar        = "═" * (w1 + 4)
    print()
    print(f"  {TG}╔{bar}╗{TRS}")
    print(f"  {TG}║{TRS}  {TW}✅  {output:<{w1}}{TG}║{TRS}")
    print(f"  {TG}║{TRS}  {TD}    {full_path:<{w1}}{TRS}{TG}║{TRS}")
    print(f"  {TG}║{TRS}  {TD}    {info_line:<{w1}}{TRS}{TG}║{TRS}")
    print(f"  {TG}╚{bar}╝{TRS}")
    print()



def main():
    print()
    print(f"{TP}  ✿ ✿ ✿ ✿ ✿ ✿ ✿ ✿ ✿ ✿ ✿ ✿ ✿ ✿ ✿ ✿ ✿ ✿ ✿ ✿ ✿{TRS}")
    print(f"{TPL}  ♡  Dedicated with love to                 ♡{TRS}")
    print()
    print(f"  {TC} ███████╗{TY}██╗  ██╗{TC}██████╗  {TY} █████╗ {TC}██╗   ██╗{TY}██╗   ██╗{TC} █████╗ {TRS}")
    print(f"  {TC} ██╔════╝{TY}██║  ██║{TC}██╔══██╗ {TY}██╔══██╗{TC}██║   ██║{TY}╚██╗ ██╔╝{TC}██╔══██╗{TRS}")
    print(f"  {TC} ███████╗{TY}███████║{TC}██████╔╝ {TY}███████║{TC}██║   ██║{TY} ╚████╔╝ {TC}███████║{TRS}")
    print(f"  {TC} ╚════██║{TY}██╔══██║{TC}██╔══██╗ {TY}██╔══██║{TC}╚██╗ ██╔╝{TY}  ╚██╔╝  {TC}██╔══██║{TRS}")
    print(f"  {TC} ███████║{TY}██║  ██║{TC}██║  ██║ {TY}██║  ██║{TC} ╚████╔╝ {TY}   ██║   {TC}██║  ██║{TRS}")
    print(f"  {TC} ╚══════╝{TY}╚═╝  ╚═╝{TC}╚═╝  ╚═╝ {TY}╚═╝  ╚═╝{TC}  ╚═══╝  {TY}   ╚═╝   {TC}╚═╝  ╚═╝{TRS}")
    print()
    print(f"{TP}                    {TW}✨ Hope you enjoy it! ✨{TRS}")
    print(f"{TP}  ✿ ✿ ✿ ✿ ✿ ✿ ✿ ✿ ✿ ✿ ✿ ✿ ✿ ✿ ✿ ✿ ✿ ✿ ✿ ✿ ✿{TRS}")
    print()
    print(f"  ╔══════════════════════════════════════════╗")
    print(f"  ║      {TW}Dictation Excel Generator{TRS}           ║")
    print(f"  ╚══════════════════════════════════════════╝")
    print(f"  {TD}Press Enter to accept defaults · Ctrl+C to cancel{TRS}")

    while True:
        run_once()
        print()
        if not ask_yes_no("Generate another file?", default=False):
            break

    print(f"\n  {TD}Goodbye!{TRS}\n")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print(f"\n\n  {TD}Cancelled.{TRS}\n")
